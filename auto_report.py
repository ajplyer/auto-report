import os
import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd
import glob

#Returns numerical value of a row based on its coffee's position in the list of coffees
#this facilitates a dynamic sorting order for the purposes of batch report formatting
def enumerate_coffee(row):
    coffee_list = open('coffee_list.txt', 'r')
    name = row[0]
    to_return = 0
    count = 0
    for line in coffee_list:
        count += 1
        if name.removesuffix(' - BULK') == line.strip():
            to_return = count
            break
        
    return to_return

#Returns string of row[1] with "Price Per Pound. . ." removed from given row to improve readability
def delete_price(row):
    contains_price = row[1]

    prefix = 4 if ' // ' in contains_price else 3
    first_half = contains_price.find('$')-prefix
    second_half = contains_price.find(',')

    first_half = contains_price[:first_half]
    second_half = contains_price[second_half:]

    return first_half+second_half

#Receives a "Coffee Of The Month" row to then be split into its corresponding coffees
#use of text file allows coffees to be updated every month or as needed
#returns list containing two tuples(rows) to be appended to excel sheet
def coffee_of_the_month(com_row, type):
    to_return = []
    if type == 'med':
        coffee = open('com_med_dark.txt')
        replaced = com_row[1].replace('Roast Level: Medium & Dark, ', '')
    elif type == 'light': 
        coffee = open('com_light.txt')
        replaced = com_row[1].replace('Roast Level: Light, ', '')
    
    replaced = 'Size: One Pound, ' + replaced

    for line in coffee:
        to_return.append((line.strip(), replaced, com_row[2]))

    return to_return

def main(file_path):

    #Nested within main() due to functionality being more centrally dependant than other functions
    #Function interprets lists from lines 123-128 and appends them to excel worksheet with appropriate formatting 
    def format_rows(current_coffee):
        current_coffee.sort(key=lambda row: row[1], reverse=True)
        

        while len(current_coffee) > 0:
            to_count = current_coffee.pop(0)
            amount = to_count[2]

            while len(current_coffee) > 0 and current_coffee[0][1] == to_count[1]:
                to_count = current_coffee.pop(0)
                amount += to_count[2]

            row = (to_count[0], to_count[1], amount)
            ws.append(row)

            if 'Five' in to_count[1] or '5' in to_count[1]:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')

            if '#0' not in to_count[1]:
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True)
    #convert batch report file downloaded by default as type .csv to the openpyxl compatible .xlsx format
    # glob.glob('*.csv')[0]
    original = pd.read_csv(file_path)
    original.to_excel(r'batch.xlsx', index=None, header=True)

    wb = openpyxl.load_workbook('batch.xlsx')

    ws = wb.active

    ws.delete_cols(4, 5)

    excel_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if 'COFFEE SUBSCRIPTION' in row[0]:
            if 'Medium & Dark' in row[0] or 'Medium & Dark' in row[1]:
                coffees = coffee_of_the_month(row, 'med')
            elif 'Light' in row[0] or 'Light' in row[1]:
                coffees = coffee_of_the_month(row, 'light')
            
            for item in coffees:
                excel_data.append((enumerate_coffee(item), item))
                
        num = enumerate_coffee(row)
        if num > 0:
            
            if '$' in row[1]:
                replacement = delete_price(row)

                new_row = (row[0], replacement, row[2])
                to_app =(num, new_row)
            else: to_app = (num, row)
        
            excel_data.append(to_app)
            
    ws.delete_rows(2, ws.max_row-1)

    excel_data.sort()

    coffee_to_sort = []
    current = 1

    # print(excel_data)

    while len(excel_data) > 0:
        coffee_to_sort.clear()

        while len(excel_data) > 0 and excel_data[0][0] == current:
            coffee = excel_data.pop(0)
            coffee_to_sort.append(coffee[1])

        format_rows(coffee_to_sort)
        
        ws.append(['blank'])
        ws['A'+str(ws.max_row)].font = Font(color='00FFFFFF')
        current += 1

    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 10

    wb.save('batch.xlsx')

if __name__ == '__main__': main()